#!/usr/bin/env python
#
# Copyright (C) 2017 - Massachusetts Institute of Technology (MIT)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

"""
Misc. Data Loading functions


Should I put the hitran data reader here as well?
i mean, after all the user should want control what data line list data
it's reading in from...

"""

import os
import sys
import json
import numpy as np
from openpyxl import load_workbook, Workbook

DIR = os.path.abspath(os.path.dirname(__file__))
sys.path.insert(0, os.path.join(DIR, '../..'))

import SEAS_Utils.common_utils.db_management2 as dbm
import SEAS_Utils.common_utils.jdx_Reader as jdx

from SEAS_Utils.common_utils.data_saver import check_path_exist, check_file_exist
#from SEAS_Main.atmosphere_effects.biosig_molecule import biosig_interpolate
from SEAS_Utils.common_utils.DIRs import molecule_info
from SEAS_Utils.common_utils.data_saver import Excel_Saver
from SEAS_Utils.common_utils.data_processor import interpolate_data


def two_column_file_loader(path,spliter=None,type="float",skip=0, skipper="#", seeker="null"):
    """
    load data from files that contains only two column
    """
    
    with open(path) as data_file:   
        data = [x.split(spliter) for x in data_file.read().split("\n")[skip:]]
        if data[-1] == []:
            data = data[:-1]        
        
        for i,info in enumerate(data):
            if skipper in info:
                continue
            else:
                data = data[i+1:]
                break
                
        for i,info in enumerate(data):
            if seeker in info:
                data = data[i+1:]
                break
        
            
            
        xdata,ydata = [list(x) for x  in zip(*data)]     
        if type == "float":
            return np.array(xdata,dtype=np.float),np.array(ydata,dtype=np.float)
        elif type == "int":
            return np.array(xdata,dtype=np.int),np.array(ydata,dtype=np.int)
        elif type == "mixed":
            return xdata,ydata
        else:
            return xdata,ydata

def two_column_chunk_file_loader(path,spliter=None,chunk_splitter="\n",type="float",skip=0):
    pass
      
def multi_column_file_loader(path,spliter=None,type="float",skip=0):
    """
    load data from files that contains multiple columns
    """
    
    with open(path) as data_file:   
        data = [x.split(spliter) for x in data_file.read().split("\n")[skip:]]
        if data[-1] == [] or data[-1] == None or data[-1] == "":
            data = data[:-1]
        
        
        data = [list(x) for x  in zip(*data)]    
        
        if type == "float":
            return np.array(data,dtype=np.float)
        elif type == "int":
            return np.array(data,dtype=np.int)
        elif type == "mixed":
            return data
    
def json_loader(path):
    
    with open(path) as data_file:
        data = json.load(data_file)

    return data

def load_npy(savepath, savename):
    
    if ".npy" in savename:
        save = os.path.join(savepath, savename)
    else:
        save = os.path.join(savepath, savename)+".npy"
        

    if check_file_exist(save):
        return np.load(save)
    else:
        print "File %s Does not exist"%save
        return []


class Excel_Loader():
    
    def __init__(self, path, name, sheetname="sheet"):
        
        #check_path_exist(path)
        #check_file_exist
        
        self.path = path
        self.name = name
        self.sheetname = sheetname
        
        self.file = os.path.join(self.path,self.name)
        
    
    def create(self):
        
        self.WBI = Workbook()
    
    
    def load(self):
        
        self.WBI = load_workbook(self.file)
        input_data = self.WBI[self.sheetname]
        

class database_loader():
    "TBD, keep this here."
    
    def __init__(self):
        
        pass
        


def molecule_cross_section_loader(inputs, cross_db, molecule):
    """
    Need to move this to the SEAS_Main
    """
    
    Pgrid = inputs["Simulation_Control"]["P_Grid"]
    Tgrid = inputs["Simulation_Control"]["T_Grid"]
    numin = inputs["Spectra"]["Numin"]
    numax = inputs["Spectra"]["Numax"]


    data = [[0 for y in range(len(Pgrid))] for x in range(len(Tgrid))]
    for i,T in enumerate(Tgrid):
        for j,P in enumerate(Pgrid):
            
            print molecule,P,T,numin,numax
            result = cross_db.c.execute("SELECT nu, coef FROM {} WHERE P={} AND T={} AND nu>={} AND nu<={} ORDER BY nu".format(molecule,P,T,numin,numax))
            fetch = np.array(result.fetchall()).T
            data[i][j] = fetch[1]
            
    nu = fetch[0]

    
    return nu, data

def molecule_cross_section_loader2(inputs, db_dir, molecule):
    """
    Need to move this to the SEAS_Main
    """

    T_grid = inputs["Simulation_Control"]["T_Grid"]
    P_grid = inputs["Simulation_Control"]["P_Grid"]

    try:

        kwargs = {"dir"        :db_dir,
                  "db_name"    :"cross_section_Simulation_%s.db"%molecule,
                  "user"       :"azariven",
                  "DEBUG"      :False,
                  "REMOVE"     :True,
                  "BACKUP"     :False,
                  "OVERWRITE"  :True}
        
        cross_db = dbm.database(**kwargs)  
        cross_db.access_db()  
    
        data = [[0 for y in range(len(T_grid))] for x in range(len(P_grid))]
        for j,P in enumerate(P_grid):
            for i,T in enumerate(T_grid):
                table_name = "T%sP%s"%(i,j)
                result = cross_db.c.execute("SELECT * FROM {} ORDER BY nu".format(table_name))
                fetch = np.array(result.fetchall()).T
                data[j][i] = fetch[1] # this is so messed up... why reverse this... why... 
                
        nu = fetch[0]
    except:
        kwargs = {"dir"        :db_dir,
                  "db_name"    :"cross_section_Simulation_%s.db"%("N2"),
                  "user"       :"azariven",
                  "DEBUG"      :False,
                  "REMOVE"     :True,
                  "BACKUP"     :False,
                  "OVERWRITE"  :True}
        
        cross_db = dbm.database(**kwargs)  
        cross_db.access_db()  
    
        data = [[0 for y in range(len(T_grid))] for x in range(len(P_grid))]
        for j,P in enumerate(P_grid):
            for i,T in enumerate(T_grid):
                table_name = "T%sP%s"%(i,j)
                result = cross_db.c.execute("SELECT * FROM {} ORDER BY nu".format(table_name))
                fetch = np.array(result.fetchall()).T
                data[j][i] = fetch[1]
                
        nu = fetch[0]

    
    return nu, data

def exomol_cross_section_loader(inputs, nu, Exomol_DB_DIR, molecule):
    
    

    T_grid = inputs["Simulation_Control"]["T_Grid"]
    P_grid = inputs["Simulation_Control"]["P_Grid"]

    #ph3_xsec = inputs["user_data"]

    """
    kwargs = {"dir"        :Exomol_DB_DIR,
              "db_name"    :"Exomol_%s.db"%molecule,
              "user"       :"azariven",
              "DEBUG"      :False,
              "REMOVE"     :True,
              "BACKUP"     :False,
              "OVERWRITE"  :True}
    
    cross_db = dbm.database(**kwargs)  
    cross_db.access_db()  
    """
    Pinit = P_grid[0]

    data = [[0 for y in range(len(T_grid))] for x in range(len(P_grid))]
    for j,P in enumerate(P_grid):
        for i,T in enumerate(T_grid):
            #table_name = "T%sP%s"%(i,j)
            #result = cross_db.c.execute("SELECT * FROM {} ORDER BY nu".format(table_name))
            #fetch = np.array(result.fetchall()).T
            
            if molecule == "NO":
                filename = "14N-16O_400-10000_%sK_1.000000.sigma"%T
            elif molecule == "PH3":
                filename = "31P-1H3_400-9999_%sK_1.000000.sigma"%T
            
            exomol_xsec = np.genfromtxt("../../input/absorption_data/Exomol_Xsec/%s/%s"%(molecule,filename),unpack=True)

            x1 = exomol_xsec[0]#np.array(fetch[0],dtype="float")
            y1 = exomol_xsec[1]#/(np.log(float(Pinit)/float(P)))#np.array(fetch[1],dtype="float")

            y2 = interpolate_data(x1,y1,nu,"C")
            data[j][i] = y2

    return nu, data


def HITRAN_Line_List_reference():
    
    from SEAS_Utils.common_utils.DIRs import HITRAN_Molecule_List
    
    molecule = open(HITRAN_Molecule_List).read().split("\n")
    
    component = []
    for i in range(len(molecule)):
        component.append([i+1,1,1])
        
    return molecule,component

def ASM_Smile_List(expect="All"):


    kwargs = {"db_name":"Molecule_DB.db",
              "user":"azariven",
              "dir":"../../input/molecule_info",
              "DEBUG":False,"REMOVE":False,"BACKUP":False,"OVERWRITE":False}
    
    cross_db = dbm.database(**kwargs)   
    cross_db.access_db()   
    
    cmd = 'SELECT SMILES, InChiKey, InChi, Formula, IUPAC_chemical_name FROM ID'
    
    result = cross_db.c.execute(cmd)
    data = np.array(result.fetchall()).T

    smiles = data[0]
    inchikeys = data[1]
    CAS = data[2]
    formula = data[3]
    name = data[4]
    
    if expect == "All":
        return smiles, inchikeys,CAS,formula,name
    elif expect == "Smiles":
        return smiles
    elif expect == "Inchikey":
        return inchikeys
    elif expect == "CAS":
        return CAS
    else:
        print "Unknown NIST output data type, Simulation Terminated"
        sys.exit()


def NIST_Smile_List(expect="All"):


    kwargs = {"db_name":"Molecule_DB.db",
              "user":"azariven",
              "dir":"../../input/molecule_info",
              "DEBUG":False,"REMOVE":False,"BACKUP":False,"OVERWRITE":False}
    
    cross_db = dbm.database(**kwargs)   
    cross_db.access_db()   
    
    cmd = 'SELECT ID.SMILES, ID.InChiKey, Spectra.CAS, ID.Formula, ID.IUPAC_chemical_name FROM ID,Spectra WHERE ID.SMILES=Spectra.Smiles AND Spectra.Is_Gas="Y"'
    
    result = cross_db.c.execute(cmd)
    data = np.array(result.fetchall()).T

    smiles = data[0]
    inchikeys = data[1]
    CAS = data[2]
    formula = data[3]
    name = data[4]
    
    if expect == "All":
        return smiles, inchikeys,CAS,formula,name
    elif expect == "Smiles":
        return smiles
    elif expect == "Inchikey":
        return inchikeys
    elif expect == "CAS":
        return CAS
    else:
        print "Unknown NIST output data type, Simulation Terminated"
        sys.exit()

def HITRAN_to_NIST(molecule, result="Smiles"):
        
    if result == "Formula":
        
        if molecule in ["C4H2", "ClO", "ClONO2", "CS", "H2", "H2O2", 
                         "HF", "HI", "HNO3", "HO2", "HOBr", "HOCl", 
                         "N2", "NO+", "NO2", "O", "O2", "OH", "SO3"]:
            print molecule, " Not in NIST DB"
            return ""
        
        elif molecule in ["C2H2", "C2H4", "C2H6", "CF4","CH3Br", "CH3Cl", "CH4", 
                          "CO", "CO2", "H2O", "H2S", "N2O",  "NO",  "O3"]:
            return molecule
        
        elif molecule in ["CH3CN","CH3OH","COF2","H2CO","HBr","HC3N","HCl","HCN",
                           "HCOOH","NH3","OCS","PH3","SF6","SO2"]:
            reference = {"CH3CN":"C2H3N", "CH3OH":"CH4O", "COF2":"CF2O", 
                         "H2CO":"CH2O","HBr":"BrH", "HC3N":"C3HN", "HCl":"ClH", 
                         "HCN":"CHN", "HCOOH":"CH2O2", "NH3":"H3N","OCS":"COS", 
                         "PH3":"H3P", "SF6":"F6S", "SO2":"O2S"}
            
            return reference[molecule] 
                          
    elif result == "Smiles":
        
        formula = HITRAN_to_NIST(molecule, result="Formula")
        
        info = NIST_Smile_List()
        
        return info[0][list(info[3]).index(formula)]
    
    else:
        print "unrecognized key, please put 'Smiles' or 'Formula'"
    

def NIST_to_HITRAN(molecule,result="Formula"): 
    
    if molecule in ["C2H2", "C2H4", "C2H6", "CF4","CH3Br", "CH3Cl", "CH4", 
                          "CO", "CO2", "H2O", "H2S", "N2O",  "NO",  "O3"]:
        return molecule
    
    elif molecule in ['F6S', 'CF2O', 'CH4O', 'O2S', 'H3N', 'H3P', 'COS', 
                      'CHN', 'C3HN', 'CH2O2', 'C2H3N', 'CH2O', 'ClH', 'BrH']:
        reference = {'COS': 'OCS', 'CH2O2': 'HCOOH', 'ClH': 'HCl', 'H3N': 'NH3', 'O2S': 'SO2', 
                     'CH4O': 'CH3OH', 'CH2O': 'H2CO', 'H3P': 'PH3', 'C2H3N': 'CH3CN', 'CHN': 'HCN', 
                     'CF2O': 'COF2', 'BrH': 'HBr', 'C3HN': 'HC3N', 'F6S': 'SF6'}
        return reference[molecule] 
    

def Particulate_Info_Loader(particulate):
  

    WB = Excel_Saver(molecule_info,"Particulate_List.xlsx")
    info = WB.load_sheet("Particulates",False)

    num = 2
    
    result = []
    while True:
        
        partical = info["A%s"%(num)].value
        common   = info["B%s"%(num)].value
        source   = info["C%s"%(num)].value
        filename = info["D%s"%(num)].value
        
        if partical == None:
            break
        if partical == particulate.lower():
            return partical, source, filename


        num+=1
    
    return None,None,None

def load_particulates(filename,output="wave"):

    if ".nc" in filename:
        
        from scipy.io import netcdf

        info = netcdf.netcdf_file(filename, 'r')
        
        text = info.variables["text"].data
        ri   = info.variables["ri"].data
        rn   = info.variables["rn"].data
        wave = info.variables["wavelength"].data
        wcm  = info.variables["wcm"].data
        lenx = info.variables["nlines"].data
        
        info.close()
    
        new_x = []
        new_rn = []
        new_ri = []
        
        
        
        if output=="wave":
            for i,dat in enumerate(wave):
                if dat <= 1:
                    continue
                
                
                if dat >= 25:
                    break
                new_x.append(wave[i])
                new_rn.append(rn[i])
                new_ri.append(ri[i])
                
        if output=="wcm":
            for i,dat in enumerate(wcm):
                if dat >= 10000:
                    continue
                
                if dat <= 500:
                    break
                new_x.append(wcm[i])
                new_rn.append(rn[i])
                new_ri.append(ri[i])
        
        
        print len(new_x)
        return new_x,new_rn,new_ri



