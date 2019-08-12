"""

table cell colors

"""



from openpyxl.styles import PatternFill
from openpyxl import load_workbook, Workbook






RedFill      = 'FF0000'
LightRedFill = 'FF6666'
FaintRedFill = 'FFCCCC'

OrangeFill      = 'FF8000'
LightOrangeFill = 'FFB266'
FaintOrangeFill = 'FFE5CC'

YellowFill      = 'FFFF00'
LightYellowFill = 'FFFF66'
FaintYellowFill = 'FFFFCC'

LimeFill        = "80FF00"
LightLimeFill   = "B2FF66"
FaintLimeFill   = "E5FFCC"

GreenFill       = "00FF00"
LightGreenFill  = "66FF66"
FaintGreenFill  = "CCFFCC"

TealFill        = "00FF80"
LightTealFill   = "66FFB2"
FaintTealFill   = "CCFFE5"

CyanFill        = "00FFFF"
LightCyanFill   = "66FFFF"
FaintCyanFill   = "CCFFFF"

BlueFill        = "0080FF"
LightBlueFill   = "66B2FF"
FaintBlueFill   = "CCE5FF"

NavyFill        = "0000FF"
LightNavyFill   = "6666FF"
FaintNavyFill   = "CCCCFF"

PurpleFill      = "7F00FF"
LightPurpleFill = "B266FF"
FaintPurpleFill = "E5CCFF"

MagentaFill     = "FF00FF"
LightMagentaFill= "FF66FF"
FaintMagentaFill= "FFCCFF"

PinkFill        = "FF007F"
LightPinkFill   = "FF66B2"
FaintPinkFill   = "FFCCE5"


def find_color(color):
    
    return PatternFill(start_color=color,
                       end_color=color,
                       fill_type = "solid")







