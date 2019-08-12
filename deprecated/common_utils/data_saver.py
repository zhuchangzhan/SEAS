#!/usr/bin/env python
#
# Copyright (C) 2017 - Massachusetts Institute of Technology (MIT)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.

"""
Misc. Data Saving functions

Code related to saving generated intermediate data
"""


import os,sys
import shutil
import numpy as np

from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Color, PatternFill

def check_file_exist(file_path):

    return os.path.isfile(file_path)
    
    
    

def check_path_exist(save_path, create=True, overwrite=False, Verbose=False):
    """
    check if path exists. If path exist and overwrite is True, will remove the existing path and content
    then replace with an empty path. If path doesn't exist and create is True, will create an empty path
    
    """
    
    if os.path.isdir(save_path):
        
        if Verbose:
            print "Path exist, ",
        if overwrite:
            shutil.rmtree(save_path)
            os.makedirs(save_path)
            if Verbose:
                print "overwrite existing path with clean path"
        else:
            if Verbose:
                print "no overwrite"
        return True
    else:
        if Verbose:
            print "Path not exist,",
        if create:
            os.makedirs(save_path)
            if Verbose:
                print "path created"
        else:
            if Verbose:
                print "path not created"        
        return False

def save_txt(savepath, savename, data, extension=".txt", overwrite=True, check=False):

    if check:
        check_path_exist(savepath)

    save = os.path.join(savepath, savename)
    
    with open(save, "w") as f:
        for i in data:
            f.write("%s %s\n"%(i[0],i[1]))
            
        f.close()
        
    
    
def save_npy(savepath, savename, data, overwrite=True, check=False):
    
    if check:
        check_path_exist(savepath)
        
    save = os.path.join(savepath, savename)

    np.save(save, data)
    
    
def save_excel(savepath, savename, data, overwrite=True):
    
    check_path_exist(savepath)
    save = os.path.join(savepath, savename)    



class Saver():
    
    
    def __init__(self, savepath, savename, data, overwrite=True):
        """
        overwrite should be on the file level
        """
        
        self.savepath = savepath
        self.savename = savename
        self.data = data
        self.overwrite = overwrite
        
        self.save = os.path.join(savepath, savename)
    
    
        self.check_data()
    
    def check_data(self):
        pass
    

    def check_file_exist(self):
    
        return os.path.isfile(self.save)
    
    def to_text(self):
        
        pass
    
    def to_npy(self):
        
        np.save(self.save, self.data)

    def to_excel(self):
        
        from openpyxl import Workbook, load_workbook
        
        pass

    
class Excel_Saver():
    
    
    yellowFill = PatternFill(start_color='FFFF00',
                   end_color='FFFF00',
                   fill_type='solid')
    redFill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    
    
    def __init__(self, path, name):

        self.path = path
        self.name = name
        
        self.filename = os.path.join(self.path,self.name)
        
        if check_file_exist(self.filename):
            self.load_workbook()
        else:
            self.create_workbook()
        
        self.input_data = None
        
        
    def create_workbook(self):
        
        self.WBI = Workbook()
        self.save()
    
    def create_worksheet(self, sheet):
        
        return self.WBI.create_sheet(title=sheet)
        
    def load_workbook(self):

        self.WBI = load_workbook(self.filename)
    
    def delete_sheet(self,sheet):
        
        self.WBI.remove_sheet(self.WBI.get_sheet_by_name(sheet))
    

    
    def load_sheet(self, sheet, default=True):
        
        if default:
            try:
                self.input_data = self.WBI[sheet]
            except:
                self.input_data = self.create_worksheet(sheet)
        else:
            try:
                return self.WBI[sheet]
            except:
                return self.create_worksheet(sheet)
            
    
    def write_column_header(self, header, offset=2):
        
        for i,cell in enumerate(header):
            
            self.input_data["A%d"%(i+offset)].value = cell        

    def write_row_header(self, header, offset=1):
        
        for i,cell in enumerate(header):
            col = chr(ord("A")+i+offset)
            self.input_data["%s%d"%(col, 1)].value = cell    
    
    def write_column(self,column_data,column_offset=0,row_offset=0):
        
        col = chr(ord("A")+column_offset)
        for i,cell in enumerate(column_data):
            self.input_data["%s%d"%(col,i+row_offset)].value = cell
    
    
    
    def write_data(self, data, row, col, color="red"):
    
        self.input_data["%s%d"%(col, row)].value = data
    
    def save(self):
        
        self.WBI.save(self.filename)
    
    
    
    
    